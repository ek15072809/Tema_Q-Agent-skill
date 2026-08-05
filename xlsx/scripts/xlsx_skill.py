"""xlsx_skill.py — Reusable building blocks for .xlsx generation.

Scope: only helpers that genuinely save lines or trap a non-obvious gotcha.
Trivial string-formatting wrappers (e.g. '=SUM(...)') are intentionally
omitted — write those formulas inline so the spreadsheet stays readable.

Provides:
  * Shared styles (FONT_HEADER, FILL_HEADER, FONT_BODY, THIN_BORDER, ...)
  * new_wb()                 — Workbook with the default sheet removed.
  * make_sheet(wb, name, headers, col_widths) — styled header + frozen panes.
  * write_row(ws, r, values) / write_rows(ws, r, rows) — styled row writers.
  * color_scale(ws, range, ...) — 3-color heatmap conditional format.
  * highlight_equal(ws, range, value, fill_hex) — single-value highlight.
  * add_dropdown(ws, range, options) — data-validation list.
  * add_bar_chart / add_line_chart — anchored charts.
  * define_name(wb, name, sheet, range) — named range.
  * save(wb, filename) — write to /home/z/my-project/download/.
"""
from __future__ import annotations
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


# ---- Shared styles --------------------------------------------------------

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill("solid", fgColor="1A1A1A")
FONT_BODY   = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


# ---- Workbook & sheets ----------------------------------------------------

def new_wb() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default "Sheet" so created sheets are clean
    return wb


def make_sheet(wb: Workbook,
               name: str,
               headers: Sequence[str],
               col_widths: Sequence[float] | None = None):
    """Create a sheet with a styled header row + frozen panes."""
    ws = wb.create_sheet(name)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    return ws


def write_row(ws, row_idx: int, values: Sequence[object]) -> None:
    """Write a styled row. Numbers right-aligned, strings left-aligned."""
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=i, value=v)
        c.font = FONT_BODY
        c.border = THIN_BORDER
        if isinstance(v, (int, float)):
            c.alignment = ALIGN_RIGHT
        else:
            c.alignment = ALIGN_LEFT


def write_rows(ws, start_row: int, rows: Iterable[Sequence[object]]) -> int:
    """Write multiple rows starting at start_row. Returns next free row."""
    r = start_row
    for row in rows:
        write_row(ws, r, row)
        r += 1
    return r


# ---- Conditional formatting ----------------------------------------------

def color_scale(ws, range_str: str,
                start_color: str = "FF6B6B",
                mid_color: str = "FFEB3B",
                end_color: str = "4CAF50") -> None:
    """3-color heatmap on a range. Colors are hex strings without #."""
    ws.conditional_formatting.add(
        range_str,
        ColorScaleRule(
            start_type="min",            start_color=start_color,
            mid_type="percentile",       mid_value=50, mid_color=mid_color,
            end_type="max",              end_color=end_color,
        ),
    )


def highlight_equal(ws, range_str: str, value: str, fill_hex: str) -> None:
    """Fill cells equal to *value* with a solid color."""
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator="equal", formula=[f'"{value}"'],
                   fill=PatternFill("solid", fgColor=fill_hex)),
    )


# ---- Data validation (dropdowns) -----------------------------------------

def add_dropdown(ws, range_str: str, options: Sequence[str],
                 error_title: str = "Input Error",
                 error_msg: str = "Invalid value") -> None:
    """Add a list-type data validation (dropdown) to a range."""
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
    )
    dv.error = error_msg
    dv.errorTitle = error_title
    ws.add_data_validation(dv)
    dv.add(range_str)


# ---- Charts ---------------------------------------------------------------
# Charts need Reference objects tied to the worksheet, so they belong in code
# rather than as a Markdown snippet.

def add_bar_chart(ws, title: str,
                  data_col: int, data_row1: int, data_row2: int,
                  cat_col: int, cat_row1: int, cat_row2: int,
                  anchor: str = "E2",
                  width: int = 18, height: int = 10) -> None:
    """Add a clustered-column bar chart anchored at *anchor*."""
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    data = Reference(ws, min_col=data_col, min_row=data_row1, max_row=data_row2)
    cats = Reference(ws, min_col=cat_col, min_row=cat_row1, max_row=cat_row2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)


def add_line_chart(ws, title: str,
                   data_col: int, data_row1: int, data_row2: int,
                   cat_col: int, cat_row1: int, cat_row2: int,
                   anchor: str = "E20") -> None:
    """Add a line chart anchored at *anchor*."""
    chart = LineChart()
    chart.title = title
    chart.style = 12
    data = Reference(ws, min_col=data_col, min_row=data_row1, max_row=data_row2)
    cats = Reference(ws, min_col=cat_col, min_row=cat_row1, max_row=cat_row2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


# ---- Defined names --------------------------------------------------------

def define_name(wb: Workbook, name: str,
                sheet: str, range_str: str) -> None:
    """Assign a readable name to a range so formulas stay readable."""
    wb.defined_names[name] = DefinedName(
        name=name, attr_text=f"'{sheet}'!{range_str}",
    )


# ---- Save -----------------------------------------------------------------

def save(wb: Workbook, filename: str,
         out_dir: str = "/home/z/my-project/download") -> Path:
    p = Path(out_dir) / filename
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    return p


# ---- Self-test ------------------------------------------------------------

if __name__ == "__main__":
    wb = new_wb()
    ws = make_sheet(wb, "Sales",
                    headers=["Region", "Q1", "Q2", "Q3", "Q4", "Total"],
                    col_widths=[14, 10, 10, 10, 10, 12])
    rows = [
        ("Tokyo",  120, 180, 240, 310, None),
        ("Osaka",   90, 140, 200, 260, None),
        ("Nagoya",  60,  90, 130, 170, None),
    ]
    next_row = write_rows(ws, 2, rows)

    # Total formula in col F — written inline because it is trivial.
    for r in range(2, next_row):
        cell = ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_RIGHT

    color_scale(ws, "E2:E4")
    add_dropdown(ws, "G2:G10", ["Tokyo", "Osaka", "Nagoya"])
    add_bar_chart(ws, "Quarterly Sales by Region",
                  data_col=2, data_row1=1, data_row2=4,
                  cat_col=1, cat_row1=2, cat_row2=4,
                  anchor="H2")
    define_name(wb, "SalesData", "Sales", "A2:F4")

    out = save(wb, "xlsx_skill_demo.xlsx")
    print(f"Saved: {out}")
